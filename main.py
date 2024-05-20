import requests
import xml.etree.ElementTree as ET
import openpyxl


def get_currency_rates():
    rates = {'RUB': 1.0}  # Добавляем рубль как базовую валюту с курсом 1
    try:
        response = requests.get('https://www.cbr.ru/scripts/XML_daily.asp')
        if response.status_code == 200:
            root = ET.fromstring(response.content)
            for child in root.findall('Valute'):
                char_code = child.find('CharCode').text
                nominal = float(child.find('Nominal').text.replace(',', '.'))
                value = float(child.find('Value').text.replace(',', '.'))
                rates[char_code] = value / nominal
        else:
            print('Ошибка получения данных')
    except Exception as e:
        print('Ошибка:', e)
    return rates

a = get_currency_rates()
file_path = 'C:/PyCh/convert/1.xlsx'

# Открытие файла
workbook = openpyxl.load_workbook(file_path)

# Получение списка всех листов
sheet_names = workbook.sheetnames
sheet = workbook.active
value1 = sheet['A2'].value
value2 = sheet['B2'].value
sheet.cell(row=2, column=3, value=float(value2) * a["EUR"])
print(value2)



workbook.save(file_path)
workbook.close()





